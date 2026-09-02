"""엑셀 가져오기/내보내기 (FR-06, FR-07).

파일 파싱은 확장자별로 openpyxl(.xlsx) / xlrd(.xls) / csv(.csv)를 사용한다.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..core import AppError, clean_str, now_dt

MAX_ROWS = 5000          # 06-3
MAX_BYTES = 10 * 1024 * 1024
EXPORT_MAX = 10000       # 07-5


# ---------------------------------------------------------------- 템플릿 정의
class Col:
    def __init__(self, header: str, field: str, required: bool = False, example: str = "", desc: str = ""):
        self.header, self.field, self.required, self.example, self.desc = header, field, required, example, desc


ASSET_COLUMNS = [
    Col("자산번호", "asset_no", True, "PC-2026-0042", "회사 고유 관리번호. 중복 불가. 영문/숫자/하이픈 30자 이내"),
    Col("자산구분", "asset_type", True, "노트북", "데스크톱 / 노트북 / 워크스테이션"),
    Col("제조사", "manufacturer", True, "삼성", "공통코드 [제조사]에 등록된 값"),
    Col("모델명", "model_name", True, "NT750XDA-KC58S", ""),
    Col("시리얼번호", "serial_no", False, "SN2024AB0091", "제조사 S/N. 값이 있으면 중복 불가"),
    Col("구매일", "purchase_date", True, "2024-03-05", "YYYY-MM-DD. 미래 날짜 불가"),
    Col("사용시작일", "service_start_date", False, "2024-03-10", "미입력 시 구매일로 간주"),
    Col("자산상태", "status", True, "사용중", "대기 / 사용중 / 수리 / 폐기예정 / 폐기"),
    Col("취득금액", "purchase_amount", False, "1450000", "숫자만. 원 단위"),
    Col("내용연수", "useful_life_years", False, "5", "연 단위 정수. 미입력 시 5년으로 간주"),
    Col("사업장", "site", True, "시화", "서울 / 인천 / 시화 / 판교 / 발안 / 음성"),
    Col("위치", "location", False, "B동 2층", "동/층 단위"),
    Col("자산관리 담당자", "manager_emp_no", True, "김IT", "해당 자산을 관리하는 IT 담당자"),
    Col("Hostname", "hostname", False, "SIH-NB-0042", "중복 시 경고(등록은 허용)"),
    Col("IP 주소", "ip_address", False, "10.20.3.41", "IPv4 형식"),
    Col("IP 구분", "ip_type", False, "고정", "고정 / DHCP"),
    Col("MAC 주소", "mac_address", False, "AC:DE:48:00:11:22", "XX:XX:XX:XX:XX:XX 또는 하이픈 형식"),
    Col("CPU", "cpu", False, "Intel i7-13700", ""),
    Col("RAM(GB)", "ram_gb", False, "16", "숫자. '16GB'로 적어도 자동 변환됨"),
    Col("디스크 유형", "disk_type", False, "NVMe", "HDD / SSD / NVMe"),
    Col("디스크 용량(GB)", "disk_gb", False, "512", "숫자. GB 단위"),
    Col("운영체제", "os", False, "Windows 11", "공통코드 [운영체제] 값"),
    Col("사번", "emp_no", False, "20210315", "상태가 '사용중'이면 필수"),
    Col("사용자명", "user_name", False, "홍길동", "상태가 '사용중'이면 필수. 사번이 마스터에 있으면 자동 보완"),
    Col("소속부서", "dept_code", False, "생산기술팀", "공통코드 [부서] 값"),
    Col("직급", "position_code", False, "대리", ""),
    Col("지급일", "issue_date", False, "2024-03-11", "상태가 '사용중'이면 필수. 구매일 이후"),
    Col("반납예정일", "due_return_date", False, "", "임대·임시 배정 시"),
    Col("폐기일", "disposal_date", False, "", "상태가 '폐기'이면 필수"),
    Col("폐기방법", "disposal_method", False, "", "상태가 '폐기'이면 필수"),
    Col("비고", "remark", False, "", ""),
]

EMPLOYEE_COLUMNS = [
    Col("사번", "emp_no", True, "20210315", "중복 불가"),
    Col("성명", "name", True, "홍길동", ""),
    Col("소속부서", "dept_code", False, "생산기술팀", "공통코드 [부서] 값"),
    Col("직급", "position_code", False, "대리", ""),
    Col("사업장", "site_code", False, "시화", ""),
    Col("재직상태", "employ_status", False, "재직", "재직 / 휴직 / 퇴사. 미입력 시 재직"),
    Col("이메일", "email", False, "", ""),
    Col("연락처", "phone", False, "", ""),
]

COLUMNS = {"asset": ASSET_COLUMNS, "employee": EMPLOYEE_COLUMNS}

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
REQ_FILL = PatternFill("solid", fgColor="2E5C8A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)


# ---------------------------------------------------------------- 템플릿 생성 (06-1)
def build_template(kind: str) -> bytes:
    cols = COLUMNS[kind]
    wb = Workbook()
    ws = wb.active
    ws.title = "데이터"

    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=c.header + (" *" if c.required else ""))
        cell.fill = REQ_FILL if c.required else HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(22, len(c.header) * 2 + 6))
    for i, c in enumerate(cols, start=1):
        ws.cell(row=2, column=i, value=c.example)
    ws.freeze_panes = "A2"

    gd = wb.create_sheet("항목 설명")
    gd.append(["컬럼", "시스템 필드", "필수", "설명"])
    for cell in gd[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for c in cols:
        gd.append([c.header, c.field, "필수" if c.required else "선택", c.desc])
    for i, w in enumerate([20, 22, 8, 60], start=1):
        gd.column_dimensions[get_column_letter(i)].width = w
    gd.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- 파일 파싱
def _norm_header(h) -> str:
    return str(h or "").replace("*", "").replace(" ", "").replace("\n", "").strip().lower()


def parse_file(filename: str, content: bytes) -> tuple[list[str], list[list]]:
    """(헤더, 데이터행) 반환. 06-2 / 06-3"""
    if len(content) > MAX_BYTES:
        raise AppError(f"파일 크기가 10MB를 초과합니다. ({len(content) / 1024 / 1024:.1f}MB)")
    name = (filename or "").lower()

    if name.endswith(".csv"):
        text = None
        for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
            try:
                text = content.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise AppError("CSV 파일의 문자 인코딩을 인식할 수 없습니다. UTF-8 또는 CP949로 저장해 주세요.")
        rows = list(csv.reader(io.StringIO(text)))
    elif name.endswith(".xls"):
        try:
            import xlrd
        except ImportError:
            raise AppError("서버에 .xls 지원 모듈이 없습니다. .xlsx로 저장 후 다시 올려주세요.")
        book = xlrd.open_workbook(file_contents=content)
        sh = book.sheet_by_index(0)
        rows = []
        for r in range(sh.nrows):
            row = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                if cell.ctype == 3:      # 날짜
                    y, mo, d, *_ = xlrd.xldate_as_tuple(cell.value, book.datemode)
                    row.append(f"{y:04d}-{mo:02d}-{d:02d}")
                elif cell.ctype == 2 and float(cell.value).is_integer():
                    row.append(int(cell.value))
                else:
                    row.append(cell.value)
            rows.append(row)
    elif name.endswith(".xlsx") or name.endswith(".xlsm"):
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        raise AppError("지원하지 않는 파일 형식입니다. .xlsx, .xls, .csv만 업로드할 수 있습니다.")

    # 완전히 빈 행 제거
    rows = [r for r in rows if any(clean_str(v) for v in r)]
    if not rows:
        raise AppError("파일에 데이터가 없습니다.")

    header = [clean_str(h) or "" for h in rows[0]]
    body = rows[1:]
    if len(body) > MAX_ROWS:
        raise AppError(f"최대 {MAX_ROWS:,}행까지 처리할 수 있습니다. (현재 {len(body):,}행)")
    if not body:
        raise AppError("헤더만 있고 데이터 행이 없습니다.")
    return header, body


def suggest_mapping(kind: str, header: list[str]) -> dict[str, str | None]:
    """엑셀 컬럼 헤더를 시스템 필드로 자동 매핑한다 (06-5의 초기값)."""
    cols = COLUMNS[kind]
    by_header = {_norm_header(c.header): c.field for c in cols}
    by_field = {c.field.lower(): c.field for c in cols}
    aliases = {
        "자산번호": "asset_no", "관리번호": "asset_no", "자산코드": "asset_no",
        "구분": "asset_type", "자산종류": "asset_type", "종류": "asset_type",
        "메이커": "manufacturer", "브랜드": "manufacturer",
        "모델": "model_name", "모델명": "model_name",
        "s/n": "serial_no", "sn": "serial_no", "시리얼": "serial_no", "일련번호": "serial_no",
        "구입일": "purchase_date", "구매일자": "purchase_date", "입고일": "purchase_date",
        "상태": "status", "자산상태": "status",
        "금액": "purchase_amount", "취득가": "purchase_amount", "구매금액": "purchase_amount",
        "근무지": "site", "사업소": "site", "사업장": "site",
        "설치위치": "location", "장소": "location",
        "담당자": "manager_emp_no", "관리자": "manager_emp_no",
        "pc명": "hostname", "컴퓨터이름": "hostname", "호스트명": "hostname",
        "ip": "ip_address", "아이피": "ip_address",
        "mac": "mac_address", "맥주소": "mac_address",
        "메모리": "ram_gb", "ram": "ram_gb", "램": "ram_gb",
        "hdd": "disk_gb", "디스크": "disk_gb", "저장용량": "disk_gb",
        "os": "os", "운영체제": "os",
        "사원번호": "emp_no", "사번": "emp_no",
        "사용자": "user_name", "사용자명": "user_name", "성명": "user_name", "이름": "user_name",
        "부서": "dept_code", "소속": "dept_code", "소속부서": "dept_code",
        "직급": "position_code", "직위": "position_code",
        "지급일": "issue_date", "배정일": "issue_date", "불출일": "issue_date",
        "반납예정": "due_return_date", "반납예정일": "due_return_date",
        "폐기일": "disposal_date", "폐기일자": "disposal_date",
        "폐기방법": "disposal_method",
        "비고": "remark", "메모": "remark",
    }
    if kind == "employee":
        aliases = {**aliases, "성명": "name", "이름": "name", "사용자명": "name",
                   "부서": "dept_code", "재직": "employ_status", "재직상태": "employ_status",
                   "근무지": "site_code", "사업장": "site_code"}

    valid = {c.field for c in cols}
    mapping: dict[str, str | None] = {}
    used: set[str] = set()
    for h in header:
        n = _norm_header(h)
        field = by_header.get(n) or by_field.get(n) or aliases.get(n)
        if field not in valid or field in used:
            field = None
        if field:
            used.add(field)
        mapping[h] = field
    return mapping


# ---------------------------------------------------------------- 오류 리포트 (06-7)
def build_error_report(errors: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "오류 리포트"
    ws.append(["행번호", "컬럼", "입력값", "오류사유"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for e in errors:
        ws.append([e.get("row"), e.get("column"), str(e.get("value") or ""), e.get("message")])
    for i, w in enumerate([10, 20, 28, 80], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{ws.max_row}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- 내보내기 (FR-07)
EXPORT_BASIC = ["asset_no", "asset_type", "manufacturer", "model_name", "status",
                "cur_user_name", "cur_emp_no", "cur_dept", "site", "location", "cur_issue_date"]

EXPORT_ALL = ["asset_no", "asset_type", "manufacturer", "model_name", "serial_no",
              "purchase_date", "service_start_date", "status", "purchase_amount",
              "useful_life_years", "site", "location", "manager_emp_no",
              "hostname", "ip_address", "ip_type", "mac_address", "cpu", "ram_gb",
              "disk_type", "disk_gb", "os", "os_eol_date",
              "cur_emp_no", "cur_user_name", "cur_dept", "cur_position",
              "cur_issue_date", "cur_due_date",
              "disposal_date", "disposal_method", "remark",
              "created_at", "created_by", "created_method", "updated_at", "updated_by"]

EXPORT_LABELS = {
    "asset_no": "자산번호", "asset_type": "자산구분", "manufacturer": "제조사",
    "model_name": "모델명", "serial_no": "시리얼번호", "purchase_date": "구매일",
    "service_start_date": "사용시작일", "status": "자산상태", "purchase_amount": "취득금액",
    "useful_life_years": "내용연수", "site": "사업장", "location": "위치",
    "manager_emp_no": "자산관리 담당자", "hostname": "Hostname", "ip_address": "IP 주소",
    "ip_type": "IP 구분", "mac_address": "MAC 주소", "cpu": "CPU", "ram_gb": "RAM(GB)",
    "disk_type": "디스크 유형", "disk_gb": "디스크 용량(GB)", "os": "운영체제",
    "os_eol_date": "OS 지원종료일", "cur_emp_no": "사번", "cur_user_name": "사용자명",
    "cur_dept": "소속부서", "cur_position": "직급", "cur_issue_date": "지급일",
    "cur_due_date": "반납예정일", "disposal_date": "폐기일", "disposal_method": "폐기방법",
    "remark": "비고", "created_at": "등록일시", "created_by": "등록자",
    "created_method": "등록방식", "updated_at": "최종수정일시", "updated_by": "최종수정자",
    # 이력 시트
    "occurred_at": "발생일시", "hist_type": "이력유형", "actor": "변경자", "reason": "사유",
    "changes": "변경 내용",
}


def _sheet(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append(r)
    for i, h in enumerate(headers, start=1):
        width = max(len(str(h)) * 2 + 4, 10)
        ws.column_dimensions[get_column_letter(i)].width = min(width, 40)
    ws.freeze_panes = "A2"                       # 07-4 틀 고정 + 자동 필터
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def build_asset_export(items: list[dict], fields: list[str],
                       history_rows: list[dict] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "자산현황"
    headers = [EXPORT_LABELS.get(f, f) for f in fields]
    rows = [[it.get(f) for f in fields] for it in items]
    _sheet(ws, headers, rows)

    if history_rows is not None:                 # 07-6 이력 포함 내보내기
        hs = wb.create_sheet("이력")
        hh = ["자산번호", "발생일시", "이력유형", "변경자", "사유", "변경 내용"]
        _sheet(hs, hh, [[h["asset_no"], h["occurred_at"], h["hist_type_label"],
                         h["actor"], h.get("reason") or "", h["summary"]] for h in history_rows])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_generic_export(title: str, headers: list[str], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:30]
    _sheet(ws, headers, rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(prefix: str = "자산현황") -> str:
    return f"{prefix}_{now_dt().strftime('%Y%m%d_%H%M')}.xlsx"     # 07-3
